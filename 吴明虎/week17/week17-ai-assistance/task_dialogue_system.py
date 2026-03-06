import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple


def levenshtein_distance(a: str, b: str) -> int:
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev_row = list(range(len(b) + 1))
    for i, ca in enumerate(a, start=1):
        curr_row = [i]
        for j, cb in enumerate(b, start=1):
            ins = curr_row[j - 1] + 1
            delete = prev_row[j] + 1
            replace = prev_row[j - 1] + (0 if ca == cb else 1)
            curr_row.append(min(ins, delete, replace))
        prev_row = curr_row
    return prev_row[-1]


def similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    dist = levenshtein_distance(a, b)
    return 1.0 - dist / max(len(a), len(b))


def split_values(raw: str) -> List[str]:
    if not raw:
        return []
    parts = re.split(r"[;,|，；、/\s]+", raw.strip())
    return [p for p in parts if p]


def read_excel_rows(excel_path: Path) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "读取 Excel 需要 openpyxl。请先执行: pip install openpyxl"
        ) from exc

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c).strip() if c is not None else "" for c in row])
    if not rows:
        return []

    header = rows[0]
    data_rows = rows[1:]
    result: List[Dict[str, str]] = []
    for row in data_rows:
        item = {}
        for i, name in enumerate(header):
            if not name:
                continue
            item[name] = row[i] if i < len(row) else ""
        result.append(item)
    return result


@dataclass
class Node:
    node_id: str
    intents: List[str]
    slots: List[str] = field(default_factory=list)
    actions: List[str] = field(default_factory=list)
    response: str = ""
    child_nodes: List[str] = field(default_factory=list)


@dataclass
class DialogueState:
    scenario_name: str
    current_node_id: Optional[str]
    active_node_id: Optional[str]
    slots: Dict[str, str] = field(default_factory=dict)
    requested_slot: Optional[str] = None
    is_finished: bool = False


class Scenario:
    def __init__(self, path: Path):
        self.path = path
        data = json.loads(path.read_text(encoding="utf-8"))
        self.nodes: Dict[str, Node] = {}
        for item in data:
            node = Node(
                node_id=item["id"],
                intents=item.get("intent", []),
                slots=item.get("slot", []),
                actions=item.get("action", []),
                response=item.get("response", ""),
                child_nodes=item.get("childnode", []),
            )
            self.nodes[node.node_id] = node
        self.root_node_id = data[0]["id"] if data else None

    def get_node(self, node_id: Optional[str]) -> Optional[Node]:
        if not node_id:
            return None
        return self.nodes.get(node_id)


class SlotOntology:
    def __init__(self):
        self.slot_questions: Dict[str, str] = {}
        self.slot_values: Dict[str, List[str]] = {}

    def load_excel(self, excel_path: Path) -> None:
        rows = read_excel_rows(excel_path)
        if not rows:
            return

        def pick_key(keys: List[str], candidates: List[str]) -> Optional[str]:
            for k in keys:
                normalized = k.strip().lower()
                for c in candidates:
                    if c in normalized:
                        return k
            return None

        keys = list(rows[0].keys())
        slot_key = pick_key(keys, ["槽", "slot"])
        ask_key = pick_key(keys, ["反问", "追问", "ask", "question"])
        value_key = pick_key(keys, ["值", "value", "枚举"])
        if not slot_key:
            raise ValueError("Excel 中未找到槽位列，请确保存在 '槽位' 或 'slot' 相关列名。")

        for row in rows:
            slot = row.get(slot_key, "").strip()
            if not slot:
                continue
            slot_name = slot if slot.startswith("#") else f"#{slot}#"
            if ask_key and row.get(ask_key, "").strip():
                self.slot_questions[slot_name] = row.get(ask_key, "").strip()
            if value_key and row.get(value_key, "").strip():
                self.slot_values[slot_name] = split_values(row.get(value_key, ""))

    def get_question(self, slot_name: str) -> str:
        if slot_name in self.slot_questions:
            return self.slot_questions[slot_name]
        pure_name = slot_name.strip("#")
        return f"请告诉我{pure_name}。"


class NLU:
    def __init__(self, ontology: SlotOntology):
        self.ontology = ontology
        self.default_values = {
            "#支付方式#": ["微信", "支付宝", "银行卡", "信用卡", "现金"],
            "#服装颜色#": ["黑", "白", "红", "蓝", "绿", "灰"],
            "#服装尺寸#": ["S", "M", "L", "XL", "XXL"],
            "#服装类型#": ["衬衫", "外套", "裤子", "卫衣", "毛衣", "T恤"],
        }
        # 困惑语句关键词
        self.confusion_keywords = ["什么", "怎么了", "是啥", "不知道", "不明白", "啥意思", "怎么回事", "为什么", "干嘛"]

    def intent_recognize(self, text: str, candidates: List[Tuple[str, str]]) -> Tuple[Optional[str], float]:
        best_node = None
        best_score = 0.0
        for node_id, intent_text in candidates:
            score = similarity(text, intent_text)
            if intent_text in text:
                score = max(score, 0.95)
            if score > best_score:
                best_score = score
                best_node = node_id
        return best_node, best_score

    def extract_slots(self, text: str, target_slots: List[str]) -> Tuple[Dict[str, str], Optional[Tuple[str, str, List[str]]]]:
        """
        提取槽位值，并验证是否在预设选项中
        返回: (提取的槽位值字典, 错误信息元组(槽位名, 无效值, 有效选项列表) 或 None)
        """
        result: Dict[str, str] = {}
        for slot in target_slots:
            value = self._extract_one_slot(text, slot)
            if value:
                # 验证提取的值是否在预设选项中
                valid, options = self.check_slot_value(slot, value)
                if not valid:
                    # 返回错误信息，但不添加到结果中
                    return result, (slot, value, options)
                result[slot] = value
        return result, None

    def normalize_free_text_slot_value(self, text: str) -> str:
        cleaned = text.strip()
        cleaned = re.sub(r"^[，。！？、,.!?\s]+|[，。！？、,.!?\s]+$", "", cleaned)
        cleaned = re.sub(r"^(我想要|我要|我想买|买|来个|来件|给我来件|给我|要)\s*", "", cleaned)
        return cleaned.strip()

    def _extract_one_slot(self, text: str, slot: str) -> Optional[str]:
        # 优先使用配置枚举值
        candidates = self.ontology.slot_values.get(slot, []) or self.default_values.get(slot, [])
        for candidate in candidates:
            if candidate and candidate in text:
                return candidate
        if candidates:
            token_candidates = re.findall(r"[\u4e00-\u9fa5A-Za-z0-9]+", text)
            best_value = None
            best_score = 0.0
            for token in token_candidates:
                for candidate in candidates:
                    score = similarity(token.lower(), candidate.lower())
                    if score > best_score:
                        best_score = score
                        best_value = candidate
            if best_score >= 0.6:
                return best_value

        # 通用规则
        if slot == "#分期付款期数#":
            m = re.search(r"(\d+)\s*期", text)
            if m:
                return m.group(1)
            m2 = re.search(r"\b(\d+)\b", text)
            if m2:
                return m2.group(1)
        if slot == "#支付方式#":
            for k in ["微信", "支付宝", "银行卡", "信用卡", "现金"]:
                if k in text:
                    return k
        if slot == "#服装尺寸#":
            m = re.search(r"\b(XXL|XL|L|M|S)\b", text, flags=re.IGNORECASE)
            if m:
                return m.group(1).upper()
            m2 = re.search(r"(\d+)\s*码", text)
            if m2:
                return f"{m2.group(1)}码"
        if slot == "#时间#":
            m = re.search(r"(\d{1,2})\s*点", text)
            if m:
                return m.group(1)
        if slot == "#电影名称#":
            m = re.search(r"看(.*?)电影", text)
            if m and m.group(1).strip():
                return m.group(1).strip()
        return None

    def is_confusion_text(self, text: str) -> bool:
        """检查用户输入是否包含困惑语句"""
        for keyword in self.confusion_keywords:
            if keyword in text:
                return True
        return False

    def check_slot_value(self, slot: str, value: str) -> Tuple[bool, List[str]]:
        """检查槽位值是否在选项列表中"""
        candidates = self.ontology.slot_values.get(slot, []) or self.default_values.get(slot, [])
        if not candidates:
            return True, []  # 没有选项限制，任何值都可以
        
        # 检查值是否在选项中
        for candidate in candidates:
            if candidate == value or candidate in value:
                return True, candidates
        
        # 检查相似度
        token_candidates = re.findall(r"[\u4e00-\u9fa5A-Za-z0-9]+", value)
        for token in token_candidates:
            for candidate in candidates:
                if similarity(token.lower(), candidate.lower()) >= 0.6:
                    return True, candidates
        
        return False, candidates


class DST:
    def update(self, state: DialogueState, recognized_slots: Dict[str, str]) -> DialogueState:
        state.slots.update(recognized_slots)
        return state

    def missing_slots(self, state: DialogueState, node: Node) -> List[str]:
        return [slot for slot in node.slots if slot not in state.slots]


class PM:
    def __init__(self, scenarios: Dict[str, Scenario], nlu: NLU, dst: DST):
        self.scenarios = scenarios
        self.nlu = nlu
        self.dst = dst

    def init_state(self, scenario_name: str) -> DialogueState:
        scenario = self.scenarios[scenario_name]
        return DialogueState(
            scenario_name=scenario_name,
            current_node_id=scenario.root_node_id,
            active_node_id=None,
            slots={},
        )

    def _candidate_nodes(self, state: DialogueState) -> List[str]:
        scenario = self.scenarios[state.scenario_name]
        current = scenario.get_node(state.current_node_id)
        if not current:
            return []
        candidates = [current.node_id]
        candidates.extend(current.child_nodes)
        return candidates

    def _get_available_functions(self, state: DialogueState, scenario: Scenario) -> str:
        """获取当前可用的功能提示"""
        # 始终显示所有主要功能，因为node1-2-3在同一层级可以相互跳转
        functions = []
        
        # 检查场景类型，根据场景显示相应的功能
        if state.scenario_name == "看电影":
            functions = [
                "开始看电影（可以说'我要看电影'）",
                "买爆米花（可以说'来个爆米花'）",
                "买可乐（可以说'来个可乐'）",
                "结束观影（可以说'结束了'或'看完了'）"
            ]
        elif state.scenario_name == "买衣服":
            functions = [
                "开始买衣服（可以说'我要买衣服'）",
                "选择分期付款（可以说'可以分期付款吗'）",
                "直接下单（可以说'我买了'）",
                "结束购物（可以说'不买了'或'退出'）"
            ]
        
        if functions:
            return "目前您可以：" + "、".join(functions) + "。"
        else:
            return "您可以开始新的对话。"

    def step(self, state: DialogueState, user_text: str) -> Tuple[DialogueState, Dict[str, str]]:
        scenario = self.scenarios[state.scenario_name]
        if state.is_finished:
            return state, {"type": "end", "text": "流程已经结束，如需继续请重新开始。"}

        # 检查用户输入是否包含困惑语句
        if self.nlu.is_confusion_text(user_text):
            # 生成困惑回复
            scenario_name = state.scenario_name
            selected_options = []
            for slot, value in state.slots.items():
                slot_name = slot.strip("#")
                selected_options.append(f"{slot_name}: {value}")
            selected_text = "，".join(selected_options) if selected_options else "暂无"
            
            current_question = ""
            if state.requested_slot:
                current_question = self.nlu.ontology.get_question(state.requested_slot)
            
            confusion_reply = f"当前在{scenario_name}场景，已经选择了：{selected_text}。"
            if current_question:
                confusion_reply += f"当前问题：{current_question}"
            return state, {"type": "confusion", "text": confusion_reply}

        candidate_node_ids = self._candidate_nodes(state)
        candidate_intents: List[Tuple[str, str]] = []
        for node_id in candidate_node_ids:
            node = scenario.get_node(node_id)
            if not node:
                continue
            for intent_text in node.intents:
                candidate_intents.append((node_id, intent_text))

        target_node_id, score = self.nlu.intent_recognize(user_text, candidate_intents)
        if target_node_id is None or score < 0.45:
            # 检查用户是否在补充槽位信息
            if state.requested_slot:
                # 用户可能在回答槽位问题，使用当前节点
                target_node_id = state.active_node_id or state.current_node_id
            else:
                # 用户输入不在预期范围内，生成拟人化提示
                available_functions = self._get_available_functions(state, scenario)
                hint_msg = f"抱歉，我没太明白您的意思。{available_functions}"
                return state, {"type": "hint", "text": hint_msg}

        node = scenario.get_node(target_node_id)
        if not node:
            return state, {"type": "fallback", "text": "抱歉，我没理解你的意思。"}

        state.active_node_id = node.node_id

        extracted, extract_error = self.nlu.extract_slots(user_text, node.slots)
        
        # 如果提取槽位值时发现无效值，立即返回错误提示
        if extract_error:
            slot, invalid_value, options = extract_error
            options_text = "、".join(options) if options else "无特定选项"
            slot_name = slot.strip("#")
            error_msg = f"抱歉，'{invalid_value}' 不是有效的{slot_name}选项。请从以下选项中选择：{options_text}"
            return state, {"type": "slot_error", "text": error_msg, "slot": slot}
        
        # 当系统正在追问某个槽位时，允许用户以自由文本直接回答，避免反复追问
        if (
            state.requested_slot
            and state.requested_slot in node.slots
            and state.requested_slot not in extracted
        ):
            fallback_value = self.nlu.normalize_free_text_slot_value(user_text)
            if fallback_value:
                # 检查槽位值是否有效
                valid, options = self.nlu.check_slot_value(state.requested_slot, fallback_value)
                if not valid:
                    # 提示用户有效的选项
                    options_text = "、".join(options) if options else "无特定选项"
                    slot_name = state.requested_slot.strip("#")
                    error_msg = f"抱歉，'{fallback_value}' 不是有效的{slot_name}选项。请从以下选项中选择：{options_text}"
                    return state, {"type": "slot_error", "text": error_msg, "slot": state.requested_slot}
                extracted[state.requested_slot] = fallback_value
        
        self.dst.update(state, extracted)
        missing = self.dst.missing_slots(state, node)
        if missing:
            state.requested_slot = missing[0]
            return state, {"type": "ask_slot", "slot": missing[0], "node_id": node.node_id}

        state.requested_slot = None
        state.current_node_id = node.node_id
        if not node.child_nodes:
            state.is_finished = True
        return state, {"type": "respond", "node_id": node.node_id}


class NLG:
    def __init__(self, ontology: SlotOntology):
        self.ontology = ontology

    def render(self, state: DialogueState, scenario: Scenario, policy_output: Dict[str, str]) -> str:
        output_type = policy_output.get("type")
        if output_type == "end":
            return policy_output.get("text", "流程结束。")
        if output_type == "fallback":
            return policy_output.get("text", "抱歉，我不太明白。")
        if output_type == "ask_slot":
            slot = policy_output["slot"]
            return self.ontology.get_question(slot)
        if output_type == "confusion":
            return policy_output.get("text", "抱歉，我不太明白你的意思。")
        if output_type == "slot_error":
            return policy_output.get("text", "抱歉，输入无效。")
        if output_type == "hint":
            return policy_output.get("text", "抱歉，我没太明白您的意思。请告诉我您想做什么。")

        node = scenario.get_node(policy_output.get("node_id"))
        if not node:
            return "抱歉，响应生成失败。"

        text = node.response
        for slot_name, slot_value in state.slots.items():
            text = text.replace(slot_name, str(slot_value))
        if node.actions:
            actions_text = "；".join(node.actions)
            return f"{text}\n[执行动作] {actions_text}"
        return text


class DialogueSystem:
    def __init__(self, scenario_dir: Path, excel_path: Optional[Path] = None):
        self.scenario_dir = scenario_dir
        self.scenarios = self._load_scenarios(scenario_dir)
        if not self.scenarios:
            raise RuntimeError(f"未在 {scenario_dir} 中找到场景 JSON。")
        self.ontology = SlotOntology()
        if excel_path and excel_path.exists():
            self.ontology.load_excel(excel_path)
        self.nlu = NLU(self.ontology)
        self.dst = DST()
        self.pm = PM(self.scenarios, self.nlu, self.dst)
        self.nlg = NLG(self.ontology)
        self.state: Optional[DialogueState] = None

    def _load_scenarios(self, scenario_dir: Path) -> Dict[str, Scenario]:
        scenarios = {}
        # 检查scenario目录是否存在
        if not scenario_dir.exists():
            # 尝试在当前目录查找
            for p in sorted(Path(".").glob("*.json")):
                scenario_name = p.stem.replace("scenario-", "")
                scenarios[scenario_name] = Scenario(p)
            # 尝试在scenario子目录查找
            scenario_subdir = Path(".") / "scenario"
            if scenario_subdir.exists():
                for p in sorted(scenario_subdir.glob("*.json")):
                    scenario_name = p.stem.replace("scenario-", "")
                    scenarios[scenario_name] = Scenario(p)
        else:
            # 正常在指定目录查找
            for p in sorted(scenario_dir.glob("*.json")):
                scenario_name = p.stem.replace("scenario-", "")
                scenarios[scenario_name] = Scenario(p)
        return scenarios

    def list_scenarios(self) -> List[str]:
        return list(self.scenarios.keys())

    def start(self, scenario_name: str) -> None:
        if scenario_name not in self.scenarios:
            raise ValueError(f"未知场景: {scenario_name}")
        self.state = self.pm.init_state(scenario_name)

    def chat(self, user_text: str) -> str:
        if not self.state:
            return "请先选择并启动一个场景。"
        state, pm_output = self.pm.step(self.state, user_text)
        self.state = state
        scenario = self.scenarios[self.state.scenario_name]
        return self.nlg.render(self.state, scenario, pm_output)


def main() -> None:
    base_dir = Path(__file__).parent
    excel_candidates = [
        base_dir / "slot_ontology.xlsx",
        base_dir / "slot_ontology.xls",
    ]
    excel_path = next((p for p in excel_candidates if p.exists()), None)
    ds = DialogueSystem(base_dir, excel_path=excel_path)

    scenarios = ds.list_scenarios()
    print("可用场景：")
    for i, name in enumerate(scenarios, start=1):
        print(f"{i}. {name}")
    choice = input("请输入场景编号: ").strip()
    if not choice.isdigit() or not (1 <= int(choice) <= len(scenarios)):
        print("无效编号，程序结束。")
        return
    scenario_name = scenarios[int(choice) - 1]
    ds.start(scenario_name)

    print(f"已进入场景：{scenario_name}")
    print("输入 quit 退出。")
    while True:
        user_text = input("你: ").strip()
        if user_text.lower() in {"quit", "exit"}:
            print("系统: 再见。")
            break
        bot = ds.chat(user_text)
        print(f"系统: {bot}")


if __name__ == "__main__":
    main()
